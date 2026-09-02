#!/usr/bin/env python3
"""
sync_avances.py — Cruza la última visita de cada IE (Sheet de visitas) con el
% de avance de obra (Matriz FFIE en SharePoint) y produce avances.json para el
dashboard.

Fuentes
  1. Visitas : CSV publicado del Google Sheet (pestaña BASE_PM)  -> VISITAS_CSV_URL
               o un archivo local .csv/.xlsx                       -> --visitas
  2. Avance  : Matriz FFIE .xlsx (SharePoint), pestaña BASE_PM      -> --matriz
               (también acepta "Nuevos o ampliados", una fila por IE)

Reglas
  - Solo se consideran filas de visitas con Mes <= mes de corte (por defecto,
    el mes en curso en America/Bogota). Los meses precargados a futuro se ignoran.
  - "Última visita" = Fecha Ejecutada más reciente por CODIGO DANE SEDE.
  - El % de avance sale de la Matriz, de la fila del MES DE LA ÚLTIMA VISITA
    EJECUTADA. Los meses precargados a futuro arrastran un % viejo y se ignoran.
  - Si una fuente falla o queda vacía, el script termina con error y NO
    sobreescribe el JSON anterior (el dashboard sigue mostrando el último dato bueno).

Uso
  python sync_avances.py --matriz matriz.xlsx --out data/avances.json
  VISITAS_CSV_URL=https://... python sync_avances.py --matriz matriz.xlsx
  python sync_avances.py --visitas matriz.xlsx --matriz matriz.xlsx   # prueba local
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timezone, timedelta
from typing import Any

import openpyxl

BOGOTA = timezone(timedelta(hours=-5))

# --------------------------------------------------------------------------- #
# Normalización
# --------------------------------------------------------------------------- #

def norm_header(s: Any) -> str:
    """'  % AVANCE EJECUTADO DE OBRA ' -> 'AVANCE EJECUTADO DE OBRA' (sin tildes)."""
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip().upper()


def norm_name(s: Any) -> str:
    """Llave de respaldo por nombre: 'I.E. San José' -> 'IE SAN JOSE'."""
    s = norm_header(s)
    s = re.sub(r"\bI\s*E(?:\s*M)?\b", "IE", s)          # I.E. / I E / IEM -> IE
    s = re.sub(r"\bINSTITUCION EDUCATIVA\b", "IE", s)
    return s


def norm_dane(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return "" if s in ("", "0", "nan", "None") else s


def parse_pct(v: Any) -> float | None:
    """0.8824 | '88.24%' | '0,00%' | 88.24 -> 88.24 (porcentaje 0-100)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        return round(x * 100, 2) if x <= 1.0 else round(x, 2)
    s = str(v).strip().replace("%", "").replace(",", ".")
    try:
        x = float(s)
    except ValueError:
        return None
    return round(x * 100, 2) if x <= 1.0 and "%" not in str(v) else round(x, 2)


DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y")


def parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().split(" ")[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_month(v: Any) -> str | None:
    """'2026-02' | datetime | 2026-02-01 -> '2026-02'."""
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return f"{v.year:04d}-{v.month:02d}"
    m = re.match(r"^\s*(\d{4})-(\d{1,2})", str(v))
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}" if m else None


# --------------------------------------------------------------------------- #
# Lectura de fuentes
# --------------------------------------------------------------------------- #

def read_table(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Convierte filas (primera = encabezado) en dicts con encabezados normalizados."""
    if not rows:
        return []
    header = [norm_header(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c not in (None, "") for c in r):
            continue
        d = {header[i]: r[i] for i in range(min(len(header), len(r))) if header[i]}
        d["__COL0"] = r[0] if r else None   # la columna de mes a veces trae encabezado "1"
        out.append(d)
    return out


def read_csv_text(text: str) -> list[dict[str, Any]]:
    return read_table(list(csv.reader(io.StringIO(text))))


def read_xlsx_sheet(path: str, sheet: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"[ERROR] La pestaña '{sheet}' no existe en {path}. Hay: {wb.sheetnames}")
    ws = wb[sheet]
    return read_table([list(r) for r in ws.iter_rows(values_only=True)])


def fetch_url(url: str) -> str:
    import urllib.request
    req = urllib.request.Request(url, headers={"User-Agent": "ffie-sync/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        raw = resp.read()
    return raw.decode("utf-8-sig")


def load_visitas(args) -> list[dict[str, Any]]:
    if args.visitas:
        if args.visitas.lower().endswith(".csv"):
            with open(args.visitas, encoding="utf-8-sig") as f:
                return read_csv_text(f.read())
        return read_xlsx_sheet(args.visitas, args.visitas_sheet)
    url = os.environ.get("VISITAS_CSV_URL")
    if not url:
        raise SystemExit("[ERROR] Falta VISITAS_CSV_URL o --visitas")
    text = fetch_url(url)
    if text.lstrip().startswith("<"):
        raise SystemExit("[ERROR] La URL de visitas devolvió HTML (¿el Sheet dejó de ser público?)")
    return read_csv_text(text)


def load_matriz(args) -> list[dict[str, Any]]:
    """Matriz de avance: .xlsx local (SharePoint descargado) o CSV publicado (MATRIZ_CSV_URL)."""
    url = os.environ.get("MATRIZ_CSV_URL")
    if args.matriz:
        if args.matriz.lower().endswith(".csv"):
            with open(args.matriz, encoding="utf-8-sig") as f:
                return read_csv_text(f.read())
        return read_xlsx_sheet(args.matriz, args.matriz_sheet)
    if not url:
        raise SystemExit("[ERROR] Falta --matriz o MATRIZ_CSV_URL")
    text = fetch_url(url)
    if text.lstrip().startswith("<"):
        raise SystemExit("[ERROR] La URL de la matriz devolvió HTML (¿dejó de ser pública?)")
    return read_csv_text(text)


# Columnas esperadas (ya normalizadas)
C_MES = ("MES", "MES YYYY MM", "1", "1 0", "__COL0")
C_DANE = "CODIGO DANE SEDE"
C_IE = "INSTITUCION EDUCATIVA"
C_SEDE = "SEDE"
C_MUN = "MUNICIPIO"
C_COORD = "COORDINACION"
C_FECHA_EJ = "FECHA EJECUTADA"
C_FECHA_PROG = "FECHA PROGRAMADA"
C_VISITA = "VISITA MES SI NO"
C_AVANCE = "AVANCE EJECUTADO DE OBRA"
C_ESTADO = "ESTADO DEL ACTA"
C_ETAPA = "ETAPA DE OBRA"
C_FASE = "FASE ACTUAL"
C_TIPO = "TIPO DE INTERVENCION"


def get(row: dict, *keys, default=None):
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return default


# --------------------------------------------------------------------------- #
# Lógica de negocio
# --------------------------------------------------------------------------- #

def ultima_visita_por_ie(visitas: list[dict], corte: str) -> tuple[dict[str, dict], list[dict]]:
    """Agrupa por DANE y devuelve la última visita ejecutada hasta el mes de corte."""
    por_ie: dict[str, dict] = {}
    sin_dane: list[dict] = []
    for r in visitas:
        mes = parse_month(get(r, *C_MES))
        if mes is None or mes > corte:
            continue
        dane = norm_dane(get(r, C_DANE))
        if not dane:
            sin_dane.append({"mes": mes, "ie": get(r, C_IE), "sede": get(r, C_SEDE)})
            continue
        fecha = parse_date(get(r, C_FECHA_EJ))
        e = por_ie.setdefault(dane, {
            "dane": dane,
            "ie": str(get(r, C_IE, default="")).strip(),
            "sede": str(get(r, C_SEDE, default="")).strip(),
            "municipio": str(get(r, C_MUN, default="")).strip(),
            "coordinacion": str(get(r, C_COORD, default="")).strip(),
            "ultima_visita": None,
            "mes_ultima_visita": None,
            "visitas_ejecutadas": 0,
            "meses_en_plan": 0,
        })
        e["meses_en_plan"] += 1
        # Los datos descriptivos más recientes ganan (por si cambió un nombre)
        for k, c in (("ie", C_IE), ("sede", C_SEDE), ("municipio", C_MUN), ("coordinacion", C_COORD)):
            v = get(r, c)
            if v:
                e[k] = str(v).strip()
        if fecha:
            e["visitas_ejecutadas"] += 1
            if e["ultima_visita"] is None or fecha > e["ultima_visita"]:
                e["ultima_visita"] = fecha
                e["mes_ultima_visita"] = mes
    return por_ie, sin_dane


def indexar_matriz(matriz: list[dict]) -> dict[str, list[dict]]:
    """
    Indexa la Matriz por DANE. Funciona con dos formas:
      - BASE_PM: una fila por IE por mes  -> lista ordenada por mes
      - Nuevos o ampliados: una fila por IE -> lista de 1 elemento sin mes
    """
    idx: dict[str, list[dict]] = {}
    for r in matriz:
        dane = norm_dane(get(r, C_DANE))
        if not dane:
            continue
        idx.setdefault(dane, []).append({
            "mes": parse_month(get(r, *C_MES)),
            "fecha_ejecutada": parse_date(get(r, C_FECHA_EJ)),
            "avance_pct": parse_pct(get(r, C_AVANCE)),
            "estado_acta": get(r, C_ESTADO),
            "etapa": get(r, C_ETAPA),
            "fase": get(r, C_FASE),
            "tipo_intervencion": get(r, C_TIPO),
            "ie_matriz": get(r, C_IE),
            "sede_matriz": get(r, C_SEDE),
        })
    for l in idx.values():
        l.sort(key=lambda x: x["mes"] or "")
    return idx


def avance_para(filas: list[dict], mes_visita: str | None, corte: str) -> dict | None:
    """
    Regla: el % de avance es el de la fila del mes de la ÚLTIMA VISITA EJECUTADA.
    Los meses precargados a futuro arrastran un % viejo, así que NO se toma
    "la fila más reciente" sino la del mes en que hubo vuelo.

    - Si esa fila no tiene %, se usa el último % no vacío en un mes <= mes_visita.
    - Si la IE no tiene visita ejecutada, se usa el último % no vacío <= corte
      (marcado como origen 'sin_visita').
    - Si la hoja es de una fila por IE (sin mes), se usa esa fila.
    """
    if not filas:
        return None
    sin_mes = [f for f in filas if f["mes"] is None]
    if sin_mes:                                   # hoja tipo "Nuevos o ampliados"
        f = sin_mes[-1]
        return {**f, "avance_origen": "matriz_actual", "avance_mes": None}

    tope = mes_visita or corte
    candidatas = [f for f in filas if f["mes"] and f["mes"] <= tope]
    if not candidatas:
        return None
    if mes_visita:
        exacta = [f for f in candidatas if f["mes"] == mes_visita and f["avance_pct"] is not None]
        if exacta:
            return {**exacta[-1], "avance_origen": "mes_ultima_visita", "avance_mes": mes_visita}
    con_valor = [f for f in candidatas if f["avance_pct"] is not None]
    if not con_valor:
        return {**candidatas[-1], "avance_origen": "sin_dato", "avance_mes": candidatas[-1]["mes"]}
    f = con_valor[-1]
    return {**f, "avance_origen": "ultimo_disponible" if mes_visita else "sin_visita", "avance_mes": f["mes"]}


def build(visitas, matriz, corte: str, ahora: datetime) -> dict:
    por_ie, sin_dane = ultima_visita_por_ie(visitas, corte)
    idx = indexar_matriz(matriz)

    ies = []
    for dane, e in por_ie.items():
        a = avance_para(idx.get(dane, []), e["mes_ultima_visita"], corte) or {}
        ies.append({
            **e,
            "ultima_visita": e["ultima_visita"].isoformat() if e["ultima_visita"] else None,
            "ie_norm": norm_name(e["ie"]),
            "sede_norm": norm_name(e["sede"]),
            "avance_pct": a.get("avance_pct"),
            "avance_mes": a.get("avance_mes"),
            "avance_origen": a.get("avance_origen"),
            "estado_acta": a.get("estado_acta"),
            "etapa": a.get("etapa"),
            "fase": a.get("fase"),
            "tipo_intervencion": a.get("tipo_intervencion"),
            "tiene_avance": a.get("avance_pct") is not None,
        })
    ies.sort(key=lambda x: (x["coordinacion"], x["municipio"], x["ie"], x["sede"]))

    solo_en_matriz = sorted(set(idx) - set(por_ie))
    sin_avance = sorted(x["dane"] for x in ies if not x["tiene_avance"])
    avances = {d: idx[d][-1] for d in idx}   # para el reporte de alertas

    return {
        "generado_en": ahora.isoformat(timespec="seconds"),
        "mes_corte": corte,
        "fuente_visitas": "Google Sheet · pestaña BASE_PM",
        "fuente_avance": "Matriz FFIE (SharePoint) · pestaña BASE_PM · % del mes de la última visita ejecutada",
        "resumen": {
            "ies_en_plan": len(ies),
            "con_avance": sum(1 for x in ies if x["tiene_avance"]),
            "sin_avance": len(sin_avance),
            "con_visita_ejecutada": sum(1 for x in ies if x["ultima_visita"]),
            "filas_sin_dane": len(sin_dane),
            "solo_en_matriz": len(solo_en_matriz),
        },
        "alertas": {
            "sin_avance_en_matriz": [{"dane": d, "ie": por_ie[d]["ie"], "sede": por_ie[d]["sede"]} for d in sin_avance],
            "solo_en_matriz": [{"dane": d, "ie": avances[d]["ie_matriz"], "sede": avances[d]["sede_matriz"]} for d in solo_en_matriz],
            "filas_sin_dane": sin_dane,
        },
        "ies": ies,
    }


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--matriz", help="Ruta al .xlsx/.csv de la Matriz FFIE. Si se omite, usa MATRIZ_CSV_URL")
    p.add_argument("--matriz-sheet", default="BASE_PM", help="Pestaña con el %% de avance (BASE_PM o 'Nuevos o ampliados')")
    p.add_argument("--visitas", help="Archivo local de visitas (.csv o .xlsx). Si se omite, usa VISITAS_CSV_URL")
    p.add_argument("--visitas-sheet", default="BASE_PM")
    p.add_argument("--corte", help="Mes de corte YYYY-MM (por defecto: mes en curso en Bogotá)")
    p.add_argument("--out", default="data/avances.json")
    args = p.parse_args()

    ahora = datetime.now(BOGOTA)
    corte = args.corte or f"{ahora.year:04d}-{ahora.month:02d}"
    if not re.match(r"^\d{4}-\d{2}$", corte):
        raise SystemExit(f"[ERROR] --corte debe ser YYYY-MM, llegó '{corte}'")

    print(f"→ Mes de corte: {corte}")
    visitas = load_visitas(args)
    print(f"→ Visitas: {len(visitas)} filas leídas")
    matriz = load_matriz(args)
    print(f"→ Matriz : {len(matriz)} filas leídas")

    if not visitas or not matriz:
        raise SystemExit("[ERROR] Una de las fuentes quedó vacía. No se actualiza el JSON.")

    data = build(visitas, matriz, corte, ahora)
    r = data["resumen"]
    if r["ies_en_plan"] == 0:
        raise SystemExit("[ERROR] 0 IEs tras el cruce. Revisa encabezados o mes de corte. No se actualiza el JSON.")

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ {args.out} escrito")
    print(f"  IEs en plan hasta {corte}: {r['ies_en_plan']}  ·  con avance: {r['con_avance']}  ·  sin avance: {r['sin_avance']}")
    print(f"  con visita ejecutada: {r['con_visita_ejecutada']}  ·  filas sin DANE: {r['filas_sin_dane']}  ·  solo en matriz: {r['solo_en_matriz']}")
    if data["alertas"]["sin_avance_en_matriz"]:
        print("  ⚠ Sin avance en la Matriz:")
        for x in data["alertas"]["sin_avance_en_matriz"]:
            print(f"     {x['dane']}  {x['ie']} — {x['sede']}")
    if data["alertas"]["filas_sin_dane"]:
        print("  ⚠ Filas de visitas sin código DANE (no se pueden cruzar):")
        for x in data["alertas"]["filas_sin_dane"]:
            print(f"     {x['mes']}  {x['ie']} — {x['sede']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
